import re
from openpyxl import load_workbook

outfile_py = "outCode.py"
remark = "### start\n"
xls_File = open("./excel_list", 'r')
xls_List = xls_File.readlines()


def funcStr(func, obj, embed_char):
    if func == "":
        str_res = embed_char[0] + obj + embed_char[1]
    else:
        str_res = func + embed_char[0] + obj + embed_char[1]
    return str_res


def objStr(obj, value, equa_char):
    split_pch = "'"
    if value == "":
        str_res = split_pch + obj + split_pch + equa_char
    else:
        str_res = split_pch + obj + split_pch + equa_char + split_pch + value + split_pch
    return str_res


def postUI(URL, parameters, values):
    for i in range(len(parameters)):
        if i != 0:
            x0 = x0 + ',' + objStr(parameters[i], values[i], ':')
        else:
            x0 = objStr(parameters[0], values[0], ':')

    x1 = funcStr("", x0, '{}')
    x2 = objStr(URL, "", ',') + x1
    x3 = funcStr('HttpSocket().execute_http_post', x2, '()')
    return URL + ' = ' + x3


def examJSON(URL, parameters, values, fid_out_py):
    json_obj = funcStr('json.loads', URL, '()') + '[0]'
    for i in range(len(parameters)):
        json_param = funcStr(json_obj, "'" + parameters[i] + "'", '[]') + "=='" + values[i] + "'"
        exam_value = funcStr('assert_true', json_param, '()')
        fid_out_py.write("\t" + exam_value + "\n")


def searchStep(work_sheet, row_step, col_step, fid_out_py):
    col_step += 1
    cell_method = work_sheet.cell(row=row_step, column=col_step).value
    col_step += 1
    # URL
    post_URL = work_sheet.cell(row=row_step, column=col_step).value
    row_step += 1
    # Params & Values
    post_parameters = []
    post_values = []
    col_parameters = col_step
    for i in range(9):
        post_param = work_sheet.cell(row=row_step, column=col_parameters).value
        post_value = work_sheet.cell(row=row_step + 1, column=col_parameters).value
        if post_param:
            post_parameters.append(post_param)
            post_values.append(post_value)
            col_parameters += 1
        else:
            break
    if cell_method == "POST":
        res = postUI(post_URL, post_parameters, post_values)
        return res, row_step + 2
    elif cell_method == "JSON":
        examJSON(post_URL, post_parameters, post_values, fid_out_py)
        return 0, row_step + 2


# --------------------------------------------------
# Generate output files
# --------------------------------------------------
# {{{
fid_out_py = open(outfile_py, "w+")
fid_out_py.write(remark + "\n")
fid_out_py.write('import numpy\n')
for xls_idx in range(len(xls_List)):
    ExcelFile = load_workbook(xls_List[xls_idx].strip())
    all_sheets = ExcelFile.sheetnames
    print("Info: Parser excel: %s!" % xls_List[xls_idx])
    for sht_idx in range(len(all_sheets)):
        if (re.match("^Expect_", all_sheets[sht_idx]) != None):  # read set_*.sheet
            print("Info: Parser sheet: %s!" % all_sheets[sht_idx])
            work_sheet = ExcelFile.worksheets[sht_idx]
            # Generate output files
            fid_out_py.write("\ndef " + work_sheet.title + "():\n")
            fid_out_py.write("\tprint(\"" + work_sheet.cell(row=1, column=1).value + "\\n\") \n")
            fid_out_py.write("\tprint(\"" + work_sheet.cell(row=2, column=1).value + "\\n\") \n")
            sheet_content = {}

            row_start = 3
            col2_start = 2
            col3_start = 3
            for i in range(work_sheet.max_row):
                if row_start > work_sheet.max_row:
                    break
                cell_value = work_sheet.cell(row=row_start, column=col2_start).value
                if cell_value == None:
                    break
                fid_out_py.write("\tprint(\"" + cell_value + ":\\n\") \n")
                row_start += 1
                if cell_value == '具体步骤':
                    for j in range(row_start, work_sheet.max_row):
                        cell_value = work_sheet.cell(row=row_start, column=col3_start).value
                        if cell_value == 'END':
                            break
                        elif cell_value == None:
                            continue
                        fid_out_py.write("\tprint(\"" + cell_value + "\\n\") \n")
                        res_step, row_start = searchStep(work_sheet, row_start, col3_start, fid_out_py)
                        fid_out_py.write("\t" + res_step + "\n")
                elif cell_value == '具体结果':
                    for j in range(row_start, work_sheet.max_row):
                        cell_value = work_sheet.cell(row=row_start, column=col3_start).value
                        if cell_value == 'END':
                            break
                        elif cell_value == None:
                            continue
                        fid_out_py.write("\tprint(\"" + cell_value + "\\n\") \n")
                        res_step, row_start = searchStep(work_sheet, row_start, col3_start, fid_out_py)
                row_start += 1

            fid_out_py.write("\tprint(\"End " + work_sheet.title + "\\n\") \n")
            fid_out_py.write("\treturn 1\n")
            fid_out_py.write("\n")
# }}}
